import keyboard
import openpyxl
import pyautogui

def openfile(INI, KEY):
    filename = "Commands.xlsx"
    ptch = sys.path[0] + "\\" + filename
    wb_obj = openpyxl.load_workbook(ptch) 
    sheet_obj = wb_obj.active 
    MaxR = sheet_obj.max_row
    MaxC = sheet_obj.max_column

    for i in range(2, MaxR):
        cell_obj1 = sheet_obj.cell(row=i, column=1) 
        if KEY in cell_obj1.internal_value:
            cell_obj2 = sheet_obj.cell(row=i, column=2) 
            if INI == 1:
                return cell_obj1.value
            if INI == 2:
                if cell_obj2.internal_value is None:
                    pyautogui.click(pyautogui.position())
                else:
                    return cell_obj2.internal_value

i = 0
while True:
    try:
        x = keyboard.read_key()
        i += 1
        if x == openfile(1, x):
            while keyboard.is_pressed(x):
                keyboard.write(openfile(2, x))
    except:
        pass
